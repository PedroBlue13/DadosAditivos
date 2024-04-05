import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog
import os

caminho = "C:\\Users\\pedro.cordeiro\\OneDrive - AF\\Área de Trabalho\\teste\\aditivos.xlsx"
caminho2 = "C:\\Users\\pedro.cordeiro\\OneDrive - AF\\Área de Trabalho\\teste\\aditivospesquisa.xlsx"


def open_block_1():
    app = ctk.CTk()
    app.geometry("550x680")
    app.title("Consulta de Equipamentos")

    sections = {
        "Propriedade": ["Aditivo", "Patrimônio", "Chamado", "Colaborador", "E-mail do colaborador"],
        "Recebimento": ["Rua", "Número", "Bairro", "Estado", "Cidade", "Responsável"],
        "Máquina": ["Tipo", "Modelo", "Armazenamento", "Processador", "Memória Ram", "Placa de Vídeo", "Fonte"],
        "Recibo e departamento": ["Número NF", "CNPJ", "Centro de custos", "Número de série", "Início do contrato", "CAMPO EXTRA"]
    }

    def create_section(app, section_title, field_names, start_row):
        section_frame = ctk.CTkFrame(app, fg_color="#E1D5E7", corner_radius=10)
        section_frame.grid(row=start_row, column=0, columnspan=4, padx=20, pady=10, sticky="we")

        title_label = ctk.CTkLabel(section_frame, text=section_title, height=25, fg_color="#3BB1A3", text_color="#FFFFFF", corner_radius=10)
        title_label.pack(side='top', fill='x', padx=10, pady=5)

        entry_grid_frame = ctk.CTkFrame(section_frame, fg_color="#E1D5E7", corner_radius=10)
        entry_grid_frame.pack(padx=10, pady=5, fill='both', expand=True)

        section = {}
        for i, field_name in enumerate(field_names):
            entry = ctk.CTkEntry(entry_grid_frame, placeholder_text=field_name, fg_color="#FFFFFF", text_color="#000000", corner_radius=10)
            entry.grid(row=i//3, column=i%3, padx=10, pady=5, sticky="we")
            entry_grid_frame.columnconfigure(i%3, weight=1)
            section[field_name] = entry

        return section

    current_row = 0
    for title, fields in sections.items():
        sections[title] = create_section(app, title, fields, current_row)
        current_row += 1

    def save_search_results(results):
        result_wb = Workbook()
        result_ws = result_wb.active

        headers = [field for fields in sections.values() for field in fields]
        result_ws.append(headers)

        for row in results:
            result_ws.append(row)

        save_path = caminho2
        result_wb.save(save_path)
        messagebox.showinfo("Sucesso", f"Resultados da pesquisa salvos em {save_path}")

    def search_data():
        try:
            wb = load_workbook(caminho)
            ws = wb.active
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado!")
            return

        sections_keys = [field for fields in sections.values() for field in fields]
        search_criteria = {}
        for section in sections.values():
            for field_name, entry_widget in section.items():
                value = entry_widget.get()
                if value:
                    search_criteria[field_name] = value

        results = []
        for row in ws.iter_rows(values_only=True):
            if all(str(row[sections_keys.index(field_name)]) == value for field_name, value in search_criteria.items()):
                results.append(row)

        if results:
            save_search_results(results)
        else:
            messagebox.showinfo("Pesquisa", "Nenhum dado correspondente encontrado.")

    search_button = ctk.CTkButton(app, text="Pesquisar", command=search_data, fg_color="#3BB1A3", text_color="white", corner_radius=10)
    search_button.grid(row=current_row, column=0, columnspan=4, padx=20, pady=20, sticky="we")

    app.mainloop()

def open_block_2():
    app = ctk.CTk()
    app.geometry("550x680") 
    app.title("Cadastro de Equipamentos")

    # Definindo as cores
    background_color = "#E1D5E7"
    section_color = "#3BB1A3"
    entry_color = "#FFFFFF"
    text_color = "#FFFFFF"  # Cor do texto nos títulos e labels
    button_color = "#3BB1A3"
    text_color_entry = "#000000"  # Cor do texto dentro dos campos de entrada
    
    app.configure(bg=background_color)
    
    # Dicionário para armazenar as seções e suas entradas de texto
    sections = {
        "Propriedade": ["Aditivo", "Patrimônio", "Chamado", "Colaborador", "E-mail do colaborador"],
        "Recebimento": ["Rua", "Número", "Bairro", "Estado", "Cidade", "Responsável"],
        "Máquina": ["Tipo", "Modelo", "Armazenamento", "Processador", "Memória Ram", "Placa de Vídeo", "Fonte"],
        "Recibo e departamento": ["Número NF", "CNPJ", "Centro de custos", "Número de série", "Início do contrato", "Observação do Pedido"]
    }
    
    def create_section(app, section_title, field_names, start_row):
        section_frame = ctk.CTkFrame(app, fg_color=background_color, corner_radius=10)
        section_frame.grid(row=start_row, column=0, columnspan=4, padx=20, pady=10, sticky="we")
    
        title_label = ctk.CTkLabel(section_frame, text=section_title, height=25, fg_color=section_color, text_color=text_color, corner_radius=10)
        title_label.pack(side='top', fill='x', padx=10, pady=5)
    
        entry_grid_frame = ctk.CTkFrame(section_frame, fg_color=background_color, corner_radius=10)
        entry_grid_frame.pack(padx=10, pady=5, fill='both', expand=True)
    
        section = {}
        for i, field_name in enumerate(field_names):
            entry = ctk.CTkEntry(entry_grid_frame, placeholder_text=field_name, fg_color=entry_color, text_color=text_color_entry, corner_radius=10)
            entry.grid(row=i//3, column=i%3, padx=10, pady=5, sticky="we")
            entry_grid_frame.columnconfigure(i%3, weight=1)
            section[field_name] = entry
    
        return section
    

    def clear_entries():
        for section in sections.values():
            for entry_widget in section.values():
                entry_widget.delete(0, 'end')

    
    def save_data(file_path):
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            # Criar cabeçalho com os nomes dos campos
            headers = []
            for field_names in sections.values():
                headers.extend(field_names)
            ws.append(headers)
    
        row_data = []
        for section in sections.values():
            for entry_widget in section.values():
                row_data.append(entry_widget.get())
        ws.append(row_data)
        wb.save(file_path)
        messagebox.showinfo("Confirmação", "Registro concluído!")
        clear_entries()
       

    current_row = 0
    for title, fields in sections.items():
        sections[title] = create_section(app, title, fields, current_row)
        current_row += 1  
    def save_data_wrapper():
        file_path = caminho
        save_data(file_path)
    
    # Botão de salvar centralizado abaixo das seções
    save_button = ctk.CTkButton(app, text="Salvar", command=save_data_wrapper, fg_color=button_color, text_color="white", corner_radius=10)
    save_button.grid(row=current_row, column=0, columnspan=4, padx=20, pady=20, sticky="we")
    
    app.mainloop()

def main():
    app = ctk.CTk()
    app.geometry("400x160")
    app.title("Dados Aditivos")

    def select_directory():
        selected_directory = filedialog.askdirectory()
        if selected_directory:
            base_path = os.path.join(selected_directory, "DadosAditivos")
            if not os.path.exists(base_path):
                os.makedirs(base_path)
            global caminho
            global caminho2
            caminho = os.path.join(base_path, "aditivos.xlsx")
            caminho2 = os.path.join(base_path, "aditivospesquisa.xlsx")
            messagebox.showinfo("Pasta Selecionada", f"Projeto baseado em: {base_path}")

    select_button = ctk.CTkButton(app, text="Selecionar Caminho", command=select_directory)
    select_button.pack(pady=10)


    button_block_1 = ctk.CTkButton(app, text="Pesquisar Aditivos", command=open_block_1)
    button_block_1.pack(pady=20)

    button_block_2 = ctk.CTkButton(app, text="Cadastrar Aditivos", command=open_block_2)
    button_block_2.pack(pady=20)

    app.mainloop()

if __name__ == "__main__":
    main()
0000000000